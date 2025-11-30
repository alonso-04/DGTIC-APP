import pandas as pd
import matplotlib.pyplot as plt
import xlwings as xw
from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from typing import Optional, List, Any

from infraestructura.dependencias import contenedor_dependencias
from utilidades.cargar_meses import cargar_mes
from configuraciones.rutas import obtener_ruta_reportes
from dominio.excepciones import ServicioValidacionError
from infraestructura.reportes.reporte_base import ReporteBase


SERVICIO_CONTROLADOR = contenedor_dependencias.obtener_servicio_controlador()


class ReporteMensualServicios(ReporteBase):
    def __init__(self):
        self.RUTA_REPORTES_SERVICOS = obtener_ruta_reportes()
    
    def crear_libro(self):
        libro = Workbook()
        return libro
    
    def crear_hojas(self, libro: Workbook):
        hoja_1 = libro.active
        hoja_1.title = "SERVICIOS PRESTADOS"
        
        hoja_2 = libro.create_sheet("REPORTES GRÁFICOS")
        
        return hoja_1, hoja_2
    
    def cargar_configuraciones_excel(self, hoja_1, hoja_2):
        hoja_1.row_dimensions[1].height = 47
    
    def cargar_encabezados_hoja_1(
        self,
        hoja_1,
        borde_celda,
        fuente_negrita,
        alineacion_centrada,
        relleno_encabezados
    ):
        encabezados = [
            "FECHA DE OFICIO RECIBIDO",
            "DIRECCIONES ADSCRITAS A LA ALCALDÍA",
            "SOLICITUD DE SERVICIO A NIVEL TÉCNICO",
            "RESPUESTA DE LA DIRECCIÓN DE INFORMÁTICA",
            "NOMBRE DEL TÉCNICO"
        ]
        
        for i, encabezado in enumerate(encabezados, start = 1):
            letra_columna = get_column_letter(i)
            
            celda = hoja_1[f"{letra_columna}1"]
            longitud_encabezado = len(encabezado)
            
            celda.border = borde_celda
            celda.font = fuente_negrita
            celda.value = encabezado
            celda.alignment = alineacion_centrada
            celda.fill = relleno_encabezados
            
            hoja_1.column_dimensions[letra_columna].width = longitud_encabezado + 18
    
    def cargar_encabezados_hoja_2(
        self,
        hoja_2,
        borde_celda,
        fuente_negrita,
        alineacion_centrada,
        relleno_encabezados
    ):
        encabezados_conteo_tipos_servicios_realizados = [
            "TIPOS DE SERVICIOS PRESTADOS",
            "CANTIDAD"
        ]
        
        encabezados_conteo_servicios_realizados_x_departamento = [
            "NOMBRE DEL DEPARTAMENTO",
            "TIPO SERVICIO PRESTADO",
            "CANTIDAD"
        ]
        
        
        # BUCLE DE ENCABEZADOS DE CONTEO DE TIPOS DE SERVICIOS REALIZADOS
        for i, encabezado in enumerate(encabezados_conteo_tipos_servicios_realizados, start = 1):
            letra_columna = get_column_letter(i)
            
            celda = hoja_2[f"{letra_columna}1"]
            longitud_encabezado = len(encabezado)
            
            celda.border = borde_celda
            celda.font = fuente_negrita
            celda.value = encabezado
            celda.alignment = alineacion_centrada
            celda.fill = relleno_encabezados
            
            if (i == 1):
                hoja_2.column_dimensions[letra_columna].width = longitud_encabezado + 25
            
            hoja_2.column_dimensions[letra_columna].width = longitud_encabezado + 10
        
        
        # BUCLE DE ENCABEZADOS DE CONTEO DE SERVICIOS REALIZADOS X DEPARTAMENTO
        for i, encabezado in enumerate(encabezados_conteo_servicios_realizados_x_departamento, start = 4):
            letra_columna = get_column_letter(i)
            
            celda = hoja_2[f"{letra_columna}1"]
            longitud_encabezado = len(encabezado)
            
            celda.border = borde_celda
            celda.font = fuente_negrita
            celda.value = encabezado
            celda.alignment = alineacion_centrada
            celda.fill = relleno_encabezados
            
            hoja_2.column_dimensions[letra_columna].width = longitud_encabezado + 10
    
    def cargar_datos_servicios_realizados(
        self,
        hoja_1,
        mes_anio,
        alineacion_centrada,
        borde_celda
    ):
        lista_dict_servicios_realizados = SERVICIO_CONTROLADOR.filtrar_mensual_servicios_controlador(mes_anio)
        datafrmae_servicios_realizados = pd.DataFrame(lista_dict_servicios_realizados)
        
        fila_actual = 2
        
        for registro in datafrmae_servicios_realizados.itertuples():
            fila = [
                registro.fecha_servicio,
                registro.nombre_departamento,
                registro.falla_presenta,
                registro.tipo_servicio_prestado,
                registro.nombres_tecnicos
            ]
            
            for indice_columna, valor_celda in enumerate(fila, start = 1):
                celda = hoja_1.cell(
                    row = fila_actual,
                    column = indice_columna,
                    value = valor_celda
                )
                
                celda.alignment = alineacion_centrada
                celda.border = borde_celda
            
            hoja_1.row_dimensions[fila_actual].height = 35
            fila_actual += 1
    
    def cargar_conteo_tipos_servicios(
        self,
        hoja_2,
        mes_anio,
        alineacion_centrada,
        borde_celda
    ):
        lista_dict_tipos_servicios_realizados = SERVICIO_CONTROLADOR.conteo_tipos_servicios_realizados_controlador(mes_anio)
        dataframe_tipos_servicios_realizados = pd.DataFrame(lista_dict_tipos_servicios_realizados)
        
        fila_actual = 2
        
        for registro in dataframe_tipos_servicios_realizados.itertuples():
            fila = [
                registro.tipo_servicio_prestado,
                registro.cantidad
            ]
            
            for indice_columna, valor_celda in enumerate(fila, start = 1):
                celda = hoja_2.cell(
                    row = fila_actual,
                    column = indice_columna,
                    value = valor_celda
                )
                
                celda.alignment = alineacion_centrada
                celda.border = borde_celda
            
            fila_actual += 1
    
    def cargar_conteo_servicios_x_departamento(
        self,
        hoja_2,
        mes_anio,
        alineacion_centrada,
        borde_celda
    ):
        lista_dict_servicios_realizados_x_departamento = SERVICIO_CONTROLADOR.conteo_servicios_realizados_x_departamento_controlador(mes_anio)
        datafrmae_servicios_realizados_x_departamento = pd.DataFrame(lista_dict_servicios_realizados_x_departamento)
        
        fila_actual = 2
        
        for registro in datafrmae_servicios_realizados_x_departamento.itertuples():
            fila = [
                registro.nombre_departamento,
                registro.tipo_servicio_prestado,
                registro.cantidad
            ]
            
            for indice_columna, valor_celda in enumerate(fila, start = 4):
                celda = hoja_2.cell(
                    row = fila_actual,
                    column = indice_columna,
                    value = valor_celda
                )
                
                celda.alignment = alineacion_centrada
                celda.border = borde_celda
            
            fila_actual += 1
    
    def cargar_grafico_conteo_tipos_servicio(self, mes_anio: str, ruta_archivo: str):
        lista_dict_tipos_servicios_realizados = SERVICIO_CONTROLADOR.conteo_tipos_servicios_realizados_controlador(mes_anio)
        dataframe_tipos_servicios_realizados = pd.DataFrame(lista_dict_tipos_servicios_realizados)
        
        figura, ax = plt.subplots(figsize = (8, 5))
        color_barras = "#54E2EE"
        
        ax.bar(
            x = dataframe_tipos_servicios_realizados["tipo_servicio_prestado"],
            width = 0.1,
            height = dataframe_tipos_servicios_realizados["cantidad"],
            color = color_barras
        )
        
        ax.set_title("CONTEO DE SERVICIOS")
        ax.tick_params(
            axis = "x",
            rotation = 45
        )
        
        figura.tight_layout()
        
        app_excel_xw = xw.App(visible = False)
        libro_xw = app_excel_xw.books.open(ruta_archivo)
        hoja_reporte_grafico = libro_xw.sheets("REPORTES GRÁFICOS")
        
        hoja_reporte_grafico.pictures.add(
            figura,
            name = "MIGRAFICO",
            top = hoja_reporte_grafico.range("H1").top,
            left = hoja_reporte_grafico.range("H1").left
        )
        
        libro_xw.save(ruta_archivo)
        libro_xw.close()
    
    def cargar_datos(self, mes_anio: str) -> Optional[List[Any]]:
        datos = []
        
        # POSICIÓN 0: MES Y AÑO
        datos.append(mes_anio)
        
        return datos
    
    def exportar(self, datos: List) -> Any:
        try:
            libro = self.crear_libro()
            hoja_1, hoja_2 = self.crear_hojas(libro)
            
            MES_ANIO = datos[0]
            
            fuente_negrita = Font(bold = True)
            relleno_encabezados = PatternFill(start_color = "00B0F0", end_color = "00B0F0", fill_type = "solid")
            alineacion_centrada = Alignment(
                horizontal = "center",
                vertical = "center",
                wrap_text = True
            )
            
            tipo_borde = Side(border_style = "thin", color = "000000")
            borde_celda = Border(
                left = tipo_borde,
                right = tipo_borde,
                top = tipo_borde,
                bottom = tipo_borde
            )
            
            anio = int(datetime.now().date().year)
            mes = cargar_mes(MES_ANIO).upper()
            nombre_archivo = f"REPORTE SERVICIOS - {mes} {anio}"
            
            self.cargar_configuraciones_excel(hoja_1, hoja_2)
            
            self.cargar_encabezados_hoja_1(
                hoja_1,
                borde_celda,
                fuente_negrita,
                alineacion_centrada,
                relleno_encabezados
            )
            
            self.cargar_encabezados_hoja_2(
                hoja_2,
                borde_celda,
                fuente_negrita,
                alineacion_centrada,
                relleno_encabezados
            )
            
            self.cargar_datos_servicios_realizados(
                hoja_1,
                MES_ANIO,
                alineacion_centrada,
                borde_celda
            )
            
            self.cargar_conteo_tipos_servicios(
                hoja_2,
                MES_ANIO,
                alineacion_centrada,
                borde_celda
            )
            
            self.cargar_conteo_servicios_x_departamento(
                hoja_2,
                MES_ANIO,
                alineacion_centrada,
                borde_celda
            )
            
            RUTA_ARCHIVO_EXCEL = f"{self.RUTA_REPORTES_SERVICOS}/{nombre_archivo}.xlsx"
            
            libro.save(RUTA_ARCHIVO_EXCEL)            
            self.cargar_grafico_conteo_tipos_servicio(MES_ANIO, RUTA_ARCHIVO_EXCEL)
        except ServicioValidacionError as error:
            raise error


if __name__ == "__main__":
    reporte_mensual_servicios = ReporteMensualServicios()
    
    try:
        MES_ANIO = "09-2025"
        datos = reporte_mensual_servicios.cargar_datos(MES_ANIO)
        reporte_mensual_servicios.exportar(datos)
    except ServicioValidacionError as error:
        print("\n".join(error.errores))