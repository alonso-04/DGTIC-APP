import pandas as pd
import matplotlib.pyplot as plt
import xlwings as xw
from datetime import date
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from typing import Optional, List, Any


from configuraciones.dependencias import contenedor_dependencias
from utilidades.cargar_meses import cargar_mes
from utilidades.tipos_reporte import TiposReporte
from configuraciones.rutas import obtener_ruta_reportes
from configuraciones.excepciones import ValidacionError, NoEncontradoError


SERVICIOS = contenedor_dependencias.obtener_servicios()
SERVICIO_TECNICO_SERVICIO = SERVICIOS["servicio_tecnico_servicio"]


class ReporteServicios:
    def crear_libro(self):
        libro = Workbook()
        return libro
    
    def crear_hojas(self, libro: Workbook):
        hoja_1 = libro.active
        hoja_1.title = "SERVICIOS PRESTADOS"
        
        hoja_2 = libro.create_sheet("REPORTES GRÁFICOS")
        
        return hoja_1, hoja_2
    
    def cargar_configuraciones_excel(self, hoja_1, hoja_2):
        hoja_1.row_dimensions[1].height = 47
    
    def cargar_encabezados_hoja_1(
        self,
        hoja_1,
        borde_celda,
        fuente_negrita,
        alineacion_centrada,
        relleno_encabezados
    ):
        encabezados = [
            "FECHA DE OFICIO RECIBIDO",
            "DIRECCIONES ADSCRITAS A LA ALCALDÍA",
            "SOLICITUD DE SERVICIO A NIVEL TÉCNICO",
            "RESPUESTA DE LA DIRECCIÓN DE INFORMÁTICA",
            "NOMBRE DEL TÉCNICO",
            "DESCRIPCIÓN",
            "CANTIDAD",
            "OBSERVACIONES"
        ]
        
        for i, encabezado in enumerate(encabezados, start = 1):
            letra_columna = get_column_letter(i)
            
            celda = hoja_1[f"{letra_columna}1"]
            longitud_encabezado = len(encabezado)
            
            celda.border = borde_celda
            celda.font = fuente_negrita
            celda.value = encabezado
            celda.alignment = alineacion_centrada
            celda.fill = relleno_encabezados
            
            hoja_1.column_dimensions[letra_columna].width = longitud_encabezado + 18
    
    def cargar_encabezados_hoja_2(
        self,
        hoja_2,
        borde_celda,
        fuente_negrita,
        alineacion_centrada,
        relleno_encabezados
    ):
        encabezados_conteo_tipos_servicios_realizados = [
            "CATEGORÍA / TIPO DE SERVICIO",
            "CANTIDAD"
        ]
        
        encabezados_conteo_servicios_realizados_x_departamento = [
            "NOMBRE DEL DEPARTAMENTO",
            "TIPO SERVICIO PRESTADO",
            "CANTIDAD"
        ]
        
        encabezados_conteo_agrupado_servicios_realizados_x_departamento = [
            "NOMBRE DEL DEPARTAMENTO",
            "CANTIDAD"
        ]
        
        
        # BUCLE DE ENCABEZADOS DE CONTEO DE TIPOS DE SERVICIOS REALIZADOS
        for i, encabezado in enumerate(encabezados_conteo_tipos_servicios_realizados, start = 1):
            letra_columna = get_column_letter(i)
            
            celda = hoja_2[f"{letra_columna}1"]
            longitud_encabezado = len(encabezado)
            
            celda.border = borde_celda
            celda.font = fuente_negrita
            celda.value = encabezado
            celda.alignment = alineacion_centrada
            celda.fill = relleno_encabezados
            
            if (i == 1):
                hoja_2.column_dimensions[letra_columna].width = longitud_encabezado + 25
            
            hoja_2.column_dimensions[letra_columna].width = longitud_encabezado + 10
        
        
        # BUCLE DE ENCABEZADOS DE CONTEO DE SERVICIOS REALIZADOS X DEPARTAMENTO
        for i, encabezado in enumerate(encabezados_conteo_servicios_realizados_x_departamento, start = 4):
            letra_columna = get_column_letter(i)
            
            celda = hoja_2[f"{letra_columna}1"]
            longitud_encabezado = len(encabezado)
            
            celda.border = borde_celda
            celda.font = fuente_negrita
            celda.value = encabezado
            celda.alignment = alineacion_centrada
            celda.fill = relleno_encabezados
            
            hoja_2.column_dimensions[letra_columna].width = longitud_encabezado + 10

        # BUCLE DE ENCABEZADOS DE CONTEO AGRUPADO DE SERVICIOS REALIZADOS X DEPARTAMENTO
        for i, encabezado in enumerate(encabezados_conteo_agrupado_servicios_realizados_x_departamento, start = 8):
            letra_columna = get_column_letter(i)
            
            celda = hoja_2[f"{letra_columna}1"]
            longitud_encabezado = len(encabezado)
            
            celda.border = borde_celda
            celda.font = fuente_negrita
            celda.value = encabezado
            celda.alignment = alineacion_centrada
            celda.fill = relleno_encabezados
            
            hoja_2.column_dimensions[letra_columna].width = longitud_encabezado + 10
    
    def cargar_datos_servicios_realizados(
        self,
        hoja_1,
        indice_opcion_tipo_reporte,
        alineacion_centrada,
        borde_celda,
        mes_anio: Optional[str] = None,
        fecha_desde: Optional[date] = None,
        fecha_hasta: Optional[date] = None,
        anio: Optional[date] = None,
        tipo_servicio_prestado: Optional[str] = None
    ):
        opcion_tipo_reporte = TiposReporte[f"{indice_opcion_tipo_reporte}"].name
        columnas = [
            "servicio_id",
            "departamento_id",
            "tipo_servicio_id",
            "nombre_departamento",
            "mes_anio",
            "fecha_servicio",
            "falla_presenta",
            "tipo_servicio_prestado",
            "nombres_tecnicos",
            "descripcion",
            "cantidad",
            "observaciones_adicionales"
        ]
        
        if (opcion_tipo_reporte == "MENSUAL"):
            tipo_reporte_mensual = SERVICIO_TECNICO_SERVICIO.obtener_tipo_reporte_a_generar(
                tipo_reporte_a_generar = opcion_tipo_reporte,
                mes_anio = mes_anio,
                tipo_servicio_prestado = tipo_servicio_prestado
            )
            lista_tupla_servicios_realizados = tipo_reporte_mensual["servicios_realizados"]
        
        if (opcion_tipo_reporte == "RANGO_FECHA"):
            tipo_reporte_rango_fecha = SERVICIO_TECNICO_SERVICIO.obtener_tipo_reporte_a_generar(
                tipo_reporte_a_generar = opcion_tipo_reporte,
                fecha_desde = fecha_desde,
                fecha_hasta = fecha_hasta,
                tipo_servicio_prestado = tipo_servicio_prestado
            )
            lista_tupla_servicios_realizados = tipo_reporte_rango_fecha["servicios_realizados"]
            
        if (opcion_tipo_reporte == "ANUAL"):
            tipo_reporte_anual = SERVICIO_TECNICO_SERVICIO.obtener_tipo_reporte_a_generar(
                tipo_reporte_a_generar = opcion_tipo_reporte,
                anio = anio,
                tipo_servicio_prestado = tipo_servicio_prestado
            )
            lista_tupla_servicios_realizados = tipo_reporte_anual["servicios_realizados"]
        
        datafrmae_servicios_realizados = pd.DataFrame(lista_tupla_servicios_realizados, columns = columnas)
        
        fila_actual = 2
        
        for registro in datafrmae_servicios_realizados.itertuples():
            fila = [
                registro.fecha_servicio,
                registro.nombre_departamento,
                registro.falla_presenta,
                registro.tipo_servicio_prestado,
                registro.nombres_tecnicos,
                registro.descripcion,
                registro.cantidad,
                registro.observaciones_adicionales
            ]
            
            for indice_columna, valor_celda in enumerate(fila, start = 1):
                celda = hoja_1.cell(
                    row = fila_actual,
                    column = indice_columna,
                    value = valor_celda
                )
                
                celda.alignment = alineacion_centrada
                celda.border = borde_celda
            
            hoja_1.row_dimensions[fila_actual].height = 35
            fila_actual += 1
    
    def cargar_conteo_tipos_servicios(
        self,
        hoja_2,
        alineacion_centrada,
        borde_celda,
        fuente_negrita,
        indice_opcion_tipo_reporte: str,
        mes_anio: Optional[str] = None,
        fecha_desde: Optional[date] = None,
        fecha_hasta: Optional[date] = None,
        anio: Optional[date] = None,
        tipo_servicio_prestado: Optional[str] = None
    ):
        opcion_tipo_reporte = TiposReporte[f"{indice_opcion_tipo_reporte}"].name
        columnas = [
            "tipo_servicio_prestado",
            "nombre_categoria",
            "total_servicios_realizados"
        ]
        
        if (opcion_tipo_reporte == "MENSUAL"):
            tipo_reporte_mensual = SERVICIO_TECNICO_SERVICIO.obtener_tipo_reporte_a_generar(
                tipo_reporte_a_generar = opcion_tipo_reporte,
                mes_anio = mes_anio,
                tipo_servicio_prestado = tipo_servicio_prestado
            )
            lista_tupla_servicios_realizados = tipo_reporte_mensual["conteo_tipos_servicio"]
        
        if (opcion_tipo_reporte == "RANGO_FECHA"):
            tipo_reporte_rango_fecha = SERVICIO_TECNICO_SERVICIO.obtener_tipo_reporte_a_generar(
                tipo_reporte_a_generar = opcion_tipo_reporte,
                fecha_desde = fecha_desde,
                fecha_hasta = fecha_hasta,
                tipo_servicio_prestado = tipo_servicio_prestado
            )
            lista_tupla_servicios_realizados = tipo_reporte_rango_fecha["conteo_tipos_servicio"]
            
        if (opcion_tipo_reporte == "ANUAL"):
            tipo_reporte_anual = SERVICIO_TECNICO_SERVICIO.obtener_tipo_reporte_a_generar(
                tipo_reporte_a_generar = opcion_tipo_reporte,
                anio = anio,
                tipo_servicio_prestado = tipo_servicio_prestado
            )
            lista_tupla_servicios_realizados = tipo_reporte_anual["conteo_tipos_servicio"]
        
        dataframe_tipos_servicios_realizados = pd.DataFrame(lista_tupla_servicios_realizados, columns = columnas)
        
        fila_actual = 2
        total_servicios_acumulados = 0
        
        categorias_agrupadas = dataframe_tipos_servicios_realizados.groupby("nombre_categoria")
        
        for nombre_categoria, grupo in categorias_agrupadas:
            celda_categoria = hoja_2.cell(row = fila_actual, column = 1, value = f"{nombre_categoria}:")
            celda_categoria.font = Font(bold = True, color = "0070C0")
            fila_actual += 1
            
            subtotal_categoria = 0
            for registro in grupo.itertuples():
                hoja_2.cell(row = fila_actual, column = 1, value = registro.tipo_servicio_prestado).alignment = Alignment(horizontal="left")
                hoja_2.cell(row = fila_actual, column = 2, value = registro.total_servicios_realizados).alignment = alineacion_centrada
                
                hoja_2.cell(row = fila_actual, column = 1).border = borde_celda
                hoja_2.cell(row = fila_actual, column = 2).border = borde_celda
                
                subtotal_categoria += registro.total_servicios_realizados
                total_servicios_acumulados += registro.total_servicios_realizados
                fila_actual += 1
            
            hoja_2.cell(row = fila_actual, column = 1, value = f"Total {nombre_categoria.title()}").font = Font(italic = True)
            celda_sub_val = hoja_2.cell(row = fila_actual, column = 2, value = subtotal_categoria)
            celda_sub_val.font = Font(bold = True)
            celda_sub_val.alignment = alineacion_centrada
            fila_actual += 2
        
        celda_total_label = hoja_2.cell(row = fila_actual, column = 1, value = "TOTAL GENERAL")
        celda_total_label.alignment = alineacion_centrada
        celda_total_label.border = borde_celda
        celda_total_label.font = fuente_negrita

        celda_total_valor = hoja_2.cell(row = fila_actual, column = 2, value = total_servicios_acumulados)
        celda_total_valor.alignment = alineacion_centrada
        celda_total_valor.border = borde_celda
        celda_total_valor.font = fuente_negrita
    
    def cargar_conteo_servicios_x_departamento(
        self,
        hoja_2,
        alineacion_centrada,
        borde_celda,
        indice_opcion_tipo_reporte: str,
        mes_anio: Optional[str] = None,
        fecha_desde: Optional[date] = None,
        fecha_hasta: Optional[date] = None,
        anio: Optional[str] = None,
        tipo_servicio_prestado: Optional[str] = None
    ):
        opcion_tipo_reporte = TiposReporte[f"{indice_opcion_tipo_reporte}"].name
        columnas = [
            "nombre_departamento",
            "tipo_servicio_prestado",
            "total_servicios_realizados"
        ]
        
        if (opcion_tipo_reporte == "MENSUAL"):
            tipo_reporte_mensual = SERVICIO_TECNICO_SERVICIO.obtener_tipo_reporte_a_generar(
                tipo_reporte_a_generar = opcion_tipo_reporte,
                mes_anio = mes_anio,
                tipo_servicio_prestado = tipo_servicio_prestado
            )
            lista_tupla_servicios_realizados = tipo_reporte_mensual["conteo_tipos_servicio_x_departamento"]
        
        if (opcion_tipo_reporte == "RANGO_FECHA"):
            tipo_reporte_rango_fecha = SERVICIO_TECNICO_SERVICIO.obtener_tipo_reporte_a_generar(
                tipo_reporte_a_generar = opcion_tipo_reporte,
                fecha_desde = fecha_desde,
                fecha_hasta = fecha_hasta,
                tipo_servicio_prestado = tipo_servicio_prestado
            )
            lista_tupla_servicios_realizados = tipo_reporte_rango_fecha["conteo_tipos_servicio_x_departamento"]
            
        if (opcion_tipo_reporte == "ANUAL"):
            tipo_reporte_anual = SERVICIO_TECNICO_SERVICIO.obtener_tipo_reporte_a_generar(
                tipo_reporte_a_generar = opcion_tipo_reporte,
                anio = anio,
                tipo_servicio_prestado = tipo_servicio_prestado
            )
            lista_tupla_servicios_realizados = tipo_reporte_anual["conteo_tipos_servicio_x_departamento"]
        
        datafrmae_servicios_realizados_x_departamento = pd.DataFrame(lista_tupla_servicios_realizados, columns = columnas)
        
        fila_actual = 2
        
        for registro in datafrmae_servicios_realizados_x_departamento.itertuples():
            fila = [
                registro.nombre_departamento,
                registro.tipo_servicio_prestado,
                registro.total_servicios_realizados
            ]
            
            for indice_columna, valor_celda in enumerate(fila, start = 4):
                celda = hoja_2.cell(
                    row = fila_actual,
                    column = indice_columna,
                    value = valor_celda
                )
                
                celda.alignment = alineacion_centrada
                celda.border = borde_celda
            
            fila_actual += 1
    
    def cargar_conteo_agrupado_servicios_x_departamento(
        self,
        hoja_2,
        alineacion_centrada,
        borde_celda,
        indice_opcion_tipo_reporte: str,
        mes_anio: Optional[str] = None,
        fecha_desde: Optional[date] = None,
        fecha_hasta: Optional[date] = None,
        anio: Optional[str] = None,
        tipo_servicio_prestado: Optional[str] = None
    ):
        opcion_tipo_reporte = TiposReporte[f"{indice_opcion_tipo_reporte}"].name
        columnas = [
            "nombre_departamento",
            "total_servicios_realizados"
        ]
        
        if (opcion_tipo_reporte == "MENSUAL"):
            tipo_reporte_mensual = SERVICIO_TECNICO_SERVICIO.obtener_tipo_reporte_a_generar(
                tipo_reporte_a_generar = opcion_tipo_reporte,
                mes_anio = mes_anio,
                tipo_servicio_prestado = tipo_servicio_prestado
            )
            lista_tupla_servicios_realizados = tipo_reporte_mensual["conteo_agrupado_servicios_x_departamento"]
        
        if (opcion_tipo_reporte == "RANGO_FECHA"):
            tipo_reporte_rango_fecha = SERVICIO_TECNICO_SERVICIO.obtener_tipo_reporte_a_generar(
                tipo_reporte_a_generar = opcion_tipo_reporte,
                fecha_desde = fecha_desde,
                fecha_hasta = fecha_hasta,
                tipo_servicio_prestado = tipo_servicio_prestado
            )
            lista_tupla_servicios_realizados = tipo_reporte_rango_fecha["conteo_agrupado_servicios_x_departamento"]
            
        if (opcion_tipo_reporte == "ANUAL"):
            tipo_reporte_anual = SERVICIO_TECNICO_SERVICIO.obtener_tipo_reporte_a_generar(
                tipo_reporte_a_generar = opcion_tipo_reporte,
                anio = anio,
                tipo_servicio_prestado = tipo_servicio_prestado
            )
            lista_tupla_servicios_realizados = tipo_reporte_anual["conteo_agrupado_servicios_x_departamento"]
        
        datafrmae_servicios_realizados_x_departamento = pd.DataFrame(lista_tupla_servicios_realizados, columns = columnas)
        
        fila_actual = 2
        
        for registro in datafrmae_servicios_realizados_x_departamento.itertuples():
            fila = [
                registro.nombre_departamento,
                registro.total_servicios_realizados
            ]
            
            for indice_columna, valor_celda in enumerate(fila, start = 8):
                celda = hoja_2.cell(
                    row = fila_actual,
                    column = indice_columna,
                    value = valor_celda
                )
                
                celda.alignment = alineacion_centrada
                celda.border = borde_celda
            
            fila_actual += 1
    
    def cargar_grafico_conteo_tipos_servicio(
        self,
        ruta_archivo: str,
        indice_opcion_tipo_reporte: str,
        mes_anio: Optional[str] = None,
        fecha_desde: Optional[date] = None,
        fecha_hasta: Optional[date] = None,
        anio: Optional[str] = None,
        tipo_servicio_prestado: Optional[str] = None
    ):
        opcion_tipo_reporte = TiposReporte[f"{indice_opcion_tipo_reporte}"].name
        columnas = [
            "tipo_servicio_prestado",
            "nombre_categoria",
            "total_servicios_realizados"
        ]
        
        if (opcion_tipo_reporte == "MENSUAL"):
            tipo_reporte_mensual = SERVICIO_TECNICO_SERVICIO.obtener_tipo_reporte_a_generar(
                tipo_reporte_a_generar = opcion_tipo_reporte,
                mes_anio = mes_anio,
                tipo_servicio_prestado = tipo_servicio_prestado
            )
            lista_tupla_servicios_realizados = tipo_reporte_mensual["conteo_tipos_servicio"]
        
        if (opcion_tipo_reporte == "RANGO_FECHA"):
            tipo_reporte_rango_fecha = SERVICIO_TECNICO_SERVICIO.obtener_tipo_reporte_a_generar(
                tipo_reporte_a_generar = opcion_tipo_reporte,
                fecha_desde = fecha_desde,
                fecha_hasta = fecha_hasta,
                tipo_servicio_prestado = tipo_servicio_prestado
            )
            lista_tupla_servicios_realizados = tipo_reporte_rango_fecha["conteo_tipos_servicio"]
            
        if (opcion_tipo_reporte == "ANUAL"):
            tipo_reporte_anual = SERVICIO_TECNICO_SERVICIO.obtener_tipo_reporte_a_generar(
                tipo_reporte_a_generar = opcion_tipo_reporte,
                anio = anio,
                tipo_servicio_prestado = tipo_servicio_prestado
            )
            lista_tupla_servicios_realizados = tipo_reporte_anual["conteo_tipos_servicio"]
        
        dataframe_tipos_servicios_realizados = pd.DataFrame(lista_tupla_servicios_realizados, columns = columnas)
        
        app_excel_xw = xw.App(visible = False)
        libro_xw = app_excel_xw.books.open(ruta_archivo)
        hoja_reporte_grafico = libro_xw.sheets("REPORTES GRÁFICOS")
        
        categorias_agrupadas = dataframe_tipos_servicios_realizados.groupby("nombre_categoria")
        columna_inicial = 11
        
        for nombre_categoria, grupo in categorias_agrupadas:
            figura, ax = plt.subplots(figsize = (6, 6))
            
            cantidad_elementos = len(grupo)
            
            if (cantidad_elementos <= 20):
                colores = plt.cm.tab20.colors[:cantidad_elementos]
            else:
                colormap = plt.colormaps.get_cmap('viridis')
                colores = [colormap(i / cantidad_elementos) for i in range(cantidad_elementos)]
            
            porciones = ax.pie(
                grupo["total_servicios_realizados"],
                labels = None,
                startangle = 90,
                colors = colores
            )[0]
            
            legendas_formateadas = [
                f"{registro.tipo_servicio_prestado}: {registro.total_servicios_realizados}"
                for registro in grupo.itertuples()
            ]
            
            ax.legend(
                porciones, 
                legendas_formateadas,
                title = "Tipos de Servicio (Cantidad)",
                loc = "upper center",
                bbox_to_anchor = (0.5, -0.05), # Desplaza la leyenda justo debajo de la torta
                ncol = 1,                      # Cambia a 2 si tienes demasiados elementos y quieres dos columnas
                fontsize = 8,
                frameon = True
            )
            
            ax.set_title(f"{nombre_categoria}", fontsize = 10, fontweight = "bold")
            ax.axis("equal")
            figura.tight_layout()
            
            letra_columna = get_column_letter(columna_inicial)
            rango_celda = f"{letra_columna}1"
            hoja_reporte_grafico.pictures.add(
                figura,
                name = f"GRAFICO_{nombre_categoria.replace(' ', '_')}",
                top = hoja_reporte_grafico.range(rango_celda).top,
                left = hoja_reporte_grafico.range(rango_celda).left
            )
            
            plt.close(figura)
            # Incrementamos 20 filas aproximadamente para que los gráficos no se encimen
            columna_inicial += 8
        
        libro_xw.save(ruta_archivo)
        libro_xw.close()
        app_excel_xw.quit()
    
    def cargar_datos(
            self,
            indice_opcion_tipo_reporte: str,
            mes_anio: Optional[str] = None,
            fecha_desde: Optional[date] = None,
            fecha_hasta: Optional[date] = None,
            anio: Optional[str] = None,
            tipo_servicio_prestado: Optional[str] = None
        ) -> Optional[List[Any]]:
        datos = []
        
        # POSICIÓN 0: ÍNDICE DE LA OPCIÓN DE TIPO DE REPORTE
        datos.append(indice_opcion_tipo_reporte)
        
        # POSICIÓN 1: MES Y AÑO
        datos.append(mes_anio)
        
        # POSICIÓN 2: FECHA DE INICIO (DESDE)
        datos.append(fecha_desde)
        
        # POSICIÓN 3: FECHA DE FIN (HASTA)
        datos.append(fecha_hasta)
        
        # POSICIÓN 4: AÑO
        datos.append(anio)
        
        # POSICIÓN 5: TIPO DE SERVICIO PRESTADO
        datos.append(tipo_servicio_prestado)
        
        return datos
    
    def exportar(self, datos: List) -> str:
        try:
            libro = self.crear_libro()
            hoja_1, hoja_2 = self.crear_hojas(libro)
            
            INDICE_OPCION_TIPO_REPORTE = datos[0]
            MES_ANIO = datos[1]
            FECHA_DESDE = datos[2]
            FECHA_HASTA = datos[3]
            ANIO = datos[4]
            TIPO_SERVICIO_PRESTADO = datos[5]
            
            fuente_negrita = Font(bold = True)
            relleno_encabezados = PatternFill(start_color = "00B0F0", end_color = "00B0F0", fill_type = "solid")
            alineacion_centrada = Alignment(
                horizontal = "center",
                vertical = "center",
                wrap_text = True
            )
            
            tipo_borde = Side(border_style = "thin", color = "000000")
            borde_celda = Border(
                left = tipo_borde,
                right = tipo_borde,
                top = tipo_borde,
                bottom = tipo_borde
            )
            
            opcion_tipo_reporte = TiposReporte[f"{INDICE_OPCION_TIPO_REPORTE}"].name
            
            if (opcion_tipo_reporte == "MENSUAL"):
                mes = cargar_mes(MES_ANIO).upper()

                _, anio = MES_ANIO.split("-")
                anio = int(anio)
                
                if (TIPO_SERVICIO_PRESTADO is not None):
                    nombre_archivo = f"REPORTE SERVICIOS - {mes} {anio} ({TIPO_SERVICIO_PRESTADO})"
                else:
                    nombre_archivo = f"REPORTE SERVICIOS - {mes} {anio}"
            
            if (opcion_tipo_reporte == "RANGO_FECHA"):
                FECHA_DESDE_FORMATEADO = FECHA_DESDE.strftime("%d-%m-%Y")
                FECHA_HASTA_FORMATEADO = FECHA_HASTA.strftime("%d-%m-%Y")
                
                if (TIPO_SERVICIO_PRESTADO is not None):
                    nombre_archivo = f"REPORTE SERVICIOS - DESDE {FECHA_DESDE_FORMATEADO} HASTA {FECHA_HASTA_FORMATEADO} ({TIPO_SERVICIO_PRESTADO})"
                else:
                    nombre_archivo = f"REPORTE SERVICIOS - DESDE {FECHA_DESDE_FORMATEADO} HASTA {FECHA_HASTA_FORMATEADO}"
            
            if (opcion_tipo_reporte == "ANUAL"):
                if (TIPO_SERVICIO_PRESTADO is not None):
                    nombre_archivo = f"REPORTE SERVICIOS - {ANIO} ({TIPO_SERVICIO_PRESTADO})"
                else:
                    nombre_archivo = f"REPORTE SERVICIOS - {ANIO}"
            
            self.cargar_configuraciones_excel(hoja_1, hoja_2)
            
            self.cargar_encabezados_hoja_1(
                hoja_1,
                borde_celda,
                fuente_negrita,
                alineacion_centrada,
                relleno_encabezados
            )
            
            self.cargar_encabezados_hoja_2(
                hoja_2,
                borde_celda,
                fuente_negrita,
                alineacion_centrada,
                relleno_encabezados
            )
            
            self.cargar_datos_servicios_realizados(
                hoja_1,
                INDICE_OPCION_TIPO_REPORTE,
                alineacion_centrada,
                borde_celda,
                MES_ANIO,
                FECHA_DESDE,
                FECHA_HASTA,
                ANIO,
                TIPO_SERVICIO_PRESTADO
            )
            
            self.cargar_conteo_tipos_servicios(
                hoja_2,
                alineacion_centrada,
                borde_celda,
                fuente_negrita,
                INDICE_OPCION_TIPO_REPORTE,
                MES_ANIO,
                FECHA_DESDE,
                FECHA_HASTA,
                ANIO,
                TIPO_SERVICIO_PRESTADO
            )
            
            self.cargar_conteo_servicios_x_departamento(
                hoja_2,
                alineacion_centrada,
                borde_celda,
                INDICE_OPCION_TIPO_REPORTE,
                MES_ANIO,
                FECHA_DESDE,
                FECHA_HASTA,
                ANIO,
                TIPO_SERVICIO_PRESTADO
            )
            
            self.cargar_conteo_agrupado_servicios_x_departamento(
                hoja_2,
                alineacion_centrada,
                borde_celda,
                INDICE_OPCION_TIPO_REPORTE,
                MES_ANIO,
                FECHA_DESDE,
                FECHA_HASTA,
                ANIO,
                TIPO_SERVICIO_PRESTADO
            )
            
            RUTA_REPORTES_SERVICOS = obtener_ruta_reportes(INDICE_OPCION_TIPO_REPORTE)
            RUTA_ARCHIVO_EXCEL = f"{RUTA_REPORTES_SERVICOS}/{nombre_archivo}.xlsx"
            
            libro.save(RUTA_ARCHIVO_EXCEL)
            
            self.cargar_grafico_conteo_tipos_servicio(
                ruta_archivo = RUTA_ARCHIVO_EXCEL,
                indice_opcion_tipo_reporte = INDICE_OPCION_TIPO_REPORTE,
                mes_anio = MES_ANIO,
                fecha_desde = FECHA_DESDE,
                fecha_hasta = FECHA_HASTA,
                anio = ANIO,
                tipo_servicio_prestado = TIPO_SERVICIO_PRESTADO
            )
            
            return RUTA_ARCHIVO_EXCEL
        except ValidacionError as error:
            raise error
        except NoEncontradoError as error:
            raise error


if __name__ == "__main__":
    reporte_servicios = ReporteServicios()
    try:
        FECHA_DESDE = date(2026, 1, 1)
        FECHA_HASTA = date(2026, 6, 30)
        MES_ANIO = "06-2026"
        datos = reporte_servicios.cargar_datos(
            indice_opcion_tipo_reporte = "RANGO_FECHA",
            fecha_desde = FECHA_DESDE,
            fecha_hasta = FECHA_HASTA
        )
        reporte_servicios.exportar(datos)
    except NoEncontradoError as error:
        print("\n".join(error.errores))
    except ValidacionError as error:
        print("\n".join(error.errores))